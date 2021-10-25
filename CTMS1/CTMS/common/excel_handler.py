# -*- coding: utf-8 -*-
""" 
@Time    : 2021/10/20 13:57
@Author  : zhx
@FileName: excel_handle.py
"""

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Worksheet

class ExcelHandler():

    """读取excel"""

    def __init__(self,file):
        """初始化函数"""
        self.file = file

    def open_sheet(self,name) -> Worksheet:
        """打开表单，在函数或者方法后面加 -> 类型：表示此函数返回值是一个这样的类型，这里是 worksheet
        函数注解
        """
        wb = load_workbook(self.file)
        sheet = wb[name]
        return sheet

    def header(self,sheet_name):
        """获取表单的表头"""
        sheet = self.open_sheet(sheet_name)
        headers = []
        for i in sheet[1]:
            headers.append(i.value)
        return headers

    def read(self,sheet_name):
        """获取所有数据"""
        sheet = self.open_sheet(sheet_name)
        # rows = list(sheet.rows)
        rows = sheet.max_row
        print('这是行数',rows)
        """ 获取标题"""
        data =[]
        for row in range(2,rows+1):
            # rows_data = []
            # for cell in row:
            #     # print('表格数据',cell)
            #     rows_data.append(cell.value)
            #     # 列表转换字典：要和 header 去zip
            #     data_dict = dict(zip(self.header(sheet_name),rows_data))
            #     # print('这是表格data数据',data_dict)
            #     # print(data_dict)
            #     data.append(data_dict)
            data_url = sheet.cell(row,4).value
            data_method = sheet.cell(row,5).value
            data_header = sheet.cell(row,6).value
            data_js_data = sheet.cell(row,7).value
            data_expected = sheet.cell(row,8).value
            a = {'url':data_url,'method':data_method,'header':data_header,'js_data':data_js_data,'expected':data_expected}
            data.append(a)
        # print('这是获取的数据',data)
        return data

    @staticmethod
    def write(file,sheet_name,row,column,data):
        """写入Excel数据"""
        wb = load_workbook(file)
        # print(wb)
        sheet = wb[sheet_name]
        # print(sheet)
        # 修改单元格
        data = sheet.cell(row,column).value
        # 保存
        wb.save(file)
        # 关闭
        wb.close()


# if __name__ == '__main__':
#     excel = ExcelHandler(r"D:/a_python/CTMS/data/cases.xlsx")
#     # 打印出excel
#     # print(excel)
#     headers = excel.header('Sheet1')
#     # 获取表单的表头，打印出表头信息
#     # print(headers)
#     reads = excel.read("Sheet1")
#     # 打印出所有数据
#     print(reads)
#     # excel.write(r'D:/a_python/CTMS/data/cases.xlsx','Sheet1',3,1,'new_value')




