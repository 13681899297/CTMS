
""""
excel的数据驱动实现：
1.excel操作流程，先操作workbook，在操作sheet，在操作cell（value)
"""
# 导入 模块
import openpyxl
# excel 操作流程
# 操作工作簿：置顶文件路径，进行文件读取，类似于open（）函数的文件读取操作
excel = openpyxl.load_workbook(r'D:\a_python\CTMS\data\cases.xlsx')
print(excel)
# 获取sheet： 基于工作簿来获取sheet页
names = excel.sheetnames
print(names)
print(type(names))
for name in names:
    print(name)
# 操作单元格cell，指定sheet页，再进行操作
sheet = excel['Sheet3']
print(sheet.values)
# 获取单元格内容
for value in sheet.values:
    print(value[0])
# 指定单元格进行内容获取
value = sheet['A5'].value
print(value)


# 获取列：最大列数
column = sheet.max_column
print(column)
# 写入前一定确保被操作文件处于关闭状态，否则会报错
# 写入：基于单元格来进行写入
sheet['H1'] = 'aa'
# 写入之后一定记得保存








